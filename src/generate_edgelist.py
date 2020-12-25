import copy
from openpyxl import load_workbook
import random
import networkx as nx
import pickle
import sys
import os.path as osp
import os
import argparse
import gc
import openpyxl

from torch import t

sys.setrecursionlimit(1000000000)
sys.path.append(os.path.realpath('.'))

from src.classes import LncRNA
from src.classes import Protein
from src.classes import LncRNA_Protein_Interaction

from src.methods import reset_basic_data, nodeSerialNumber_listIndex_dict_generation, nodeName_listIndex_dict_generation


def parse_args():
    parser = argparse.ArgumentParser(description="generate_dataset.")
    parser.add_argument('--projectName',  help='project name')
    parser.add_argument('--interactionDatasetName', default="NPInter2", help='raw interactions dataset')
    parser.add_argument('--createBalanceDataset', default=1, type=int, help='Create a Balance dataset')
    parser.add_argument('--reduce', default=0, type=int, help='randomly reduce the source database, and also maintain one connected component')
    parser.add_argument('--reduceRatio', default=0.5, help='reduce Ratio')
    parser.add_argument('--output', default=0,type=int,  help='output dataset or not')

    return parser.parse_args()


def read_interaction_dataset(dataset_path, dataset_name):
    interaction_list = []
    negative_interaction_list = []
    lncRNA_list = []
    protein_list = []
    lncRNA_name_index_dict = {}
    protein_name_index_dict = {}
    set_interactionKey = set()
    set_negativeInteractionKey = set()
    # lncRNA_name_index_dict, protein_name_index_dic为了在interaction dataset中，读到重复的lncRNA或protein时
    # 能在lncRNA_list和protein_list中快速的找到
    if not osp.exists(dataset_path):
        raise Exception('interaction datset does not exist')
    print('start reading xlsx file')
    wb = load_workbook(dataset_path)
    sheets = wb.worksheets   # 获取当前所有的sheet
    sheet = sheets[0]
    rows = sheet.rows

    serial_number = 0
    lncRNA_count = 0
    protein_count = 0
    flag = 0 #用来排除RPI7317.xlsx的第一行

    for row in rows:
        #排除第一行
        if flag == 0:
            flag = flag + 1
            continue
        #读出这一行的每个元素,每一行对应一个interaction实例，如果这个interaction对应的lncRNA和protein还没创建，就创建它
        #并在索引词典中加入它在lncRNA_list或者protein_list中的索引
        [lncRNA_name, protein_name, label] = [col.value for col in row]
        label = int(label)
        if lncRNA_name not in lncRNA_name_index_dict:   # 新的，没创建过的lncRNA
            temp_lncRNA = LncRNA(lncRNA_name, serial_number, 'LncRNA')
            lncRNA_list.append(temp_lncRNA)
            lncRNA_name_index_dict[lncRNA_name] = lncRNA_count
            serial_number = serial_number + 1
            lncRNA_count = lncRNA_count + 1
        else:   # 在interaction dataset中已经读到过，已经创建了对象的lncRNA，就存在lncRNA_list中
            temp_lncRNA = lncRNA_list[lncRNA_name_index_dict[lncRNA_name]]
        if protein_name not in protein_name_index_dict: # 新的，没创建过的protein
            temp_protein = Protein(protein_name, serial_number, 'Protein')
            protein_list.append(temp_protein)
            protein_name_index_dict[protein_name] = protein_count
            serial_number = serial_number + 1
            protein_count = protein_count + 1
        else:   # 在interaction dataset中已经读到过，已经创建了对象的protein，就存在protein_list中
            temp_protein = protein_list[protein_name_index_dict[protein_name]]
        # 创建新的相互作用实例
        interaction_key = (temp_lncRNA.serial_number, temp_protein.serial_number)
        temp_interaction = LncRNA_Protein_Interaction(temp_lncRNA, temp_protein, label, interaction_key)
        temp_lncRNA.interaction_list.append(temp_interaction)
        temp_protein.interaction_list.append(temp_interaction)
        # print(temp_interaction.protein.name, temp_interaction.lncRNA.name)

        if label == 1:
            interaction_list.append(temp_interaction)
            set_interactionKey.add(interaction_key)
        elif label == 0:
            negative_interaction_list.append(temp_interaction)
            set_negativeInteractionKey.add(interaction_key)
        else:
            print(label)
            raise Exception('{dataset_name}has labels other than 0 and 1'.format(dataset_name=dataset_name))

    print('number of lncRNA：{:d}, number of protein：{:d}, number of node：{:d}'.format(lncRNA_count, protein_count, lncRNA_count + protein_count))
    print('number of interaction：{:d}'.format(len(interaction_list) + len(negative_interaction_list)))
    return interaction_list, negative_interaction_list,lncRNA_list, protein_list, lncRNA_name_index_dict, protein_name_index_dict, set_interactionKey, set_negativeInteractionKey


def negative_interaction_generation():
    global lncRNA_list, protein_list, interaction_list, negative_interaction_list, set_interactionKey, set_negativeInteractionKey
    set_negativeInteractionKey = set()

    if len(negative_interaction_list) != 0:
        raise Exception('negative interactions exist')

    num_of_interaction = len(interaction_list)
    num_of_lncRNA = len(lncRNA_list)
    num_of_protein = len(protein_list)

    negative_interaction_count = 0
    while(negative_interaction_count < num_of_interaction):
        random_index_lncRNA = random.randint(0, num_of_lncRNA - 1)
        random_index_protein = random.randint(0, num_of_protein - 1)
        temp_lncRNA = lncRNA_list[random_index_lncRNA]
        temp_protein = protein_list[random_index_protein]
        # 检查随机选出的lncRNA和protein是不是有已知相互作用
        key_negativeInteraction = (temp_lncRNA.serial_number, temp_protein.serial_number)
        if key_negativeInteraction in set_interactionKey:
            continue
        if key_negativeInteraction in set_negativeInteractionKey:
            continue

        # 经过检查，随机选出的lncRNA和protein是可以作为负样本的
        set_negativeInteractionKey.add(key_negativeInteraction)
        temp_interaction = LncRNA_Protein_Interaction(temp_lncRNA, temp_protein, 0, key_negativeInteraction)
        negative_interaction_list.append(temp_interaction)
        temp_lncRNA.interaction_list.append(temp_interaction)
        temp_protein.interaction_list.append(temp_interaction)
        negative_interaction_count = negative_interaction_count + 1
    print('generate ', len(negative_interaction_list), ' negative samples')


def return_node_list_and_edge_list():
    global interaction_list, negative_interaction_list, lncRNA_list, protein_list

    node_list = lncRNA_list[:]
    node_list.extend(protein_list)
    edge_list = interaction_list[:]
    edge_list.extend(negative_interaction_list)
    return node_list, edge_list


# def read_node2vec_result(path):
#     print('读入，node2vec的结果')
#     node_list, edge_list = return_node_list_and_edge_list()
#     serialNumber_listIndex_dict = nodeSerialNumber_listIndex_dict_generation(node_list)
#
#     node2vec_result_file = open(path, mode='r')
#     lines = node2vec_result_file.readlines()
#     lines.pop(0)    # 第一行包含：节点个数 节点嵌入后的维度
#     for line in lines:
#         arr = line.strip().split(' ')
#         serial_number = int(arr[0])
#         arr.pop(0)
#         node_list[serialNumber_listIndex_dict[serial_number]].embedded_vector = arr
#
#     for node in node_list:
#         if len(node.embedded_vector) != 64:
#             print('node2vec读入有问题')
#     node2vec_result_file.close()


# def load_node_k_mer(node_list, node_type, k_mer_path):
#     node_name_index_dict = nodeName_listIndex_dict_generation(node_list)   # 节点的名字：节点在其所在的列表中的index
#     with open(k_mer_path, mode='r') as f:   # 打开存有k-mer特征的文件
#         lines = f.readlines()
#         # 读取k-mer文件
#         for i in range(len(lines)):
#             line = lines[i]
#             # 从文件中定位出lncRNA或者protein的名字
#             if line[0] == '>':
#                 node_name = line.strip()[1:]
#                 if node_name in node_name_index_dict:   # 根据名字在node_list中找到它，把k-mer数据赋予它
#                     node = node_list[node_name_index_dict[node_name]]
#                     if len(node.attributes_vector) == 0:    # 如果一个node的attributes_vector已经被赋值过，不重复赋值
#                     # 如果这个node，已经被赋予过k-mer数据，报出异常
#                         if len(node.attributes_vector) != 0:
#                             print(node_name, node.node_type)
#                             raise Exception('node被赋予过k-mer数据')
#                         # k-mer数据提取出来，根据node是lncRNA还是protein，给attributes_vector赋值
#                         k_mer_vector = lines[i + 1].strip().split('\t')
#                         if node_type == 'lncRNA':
#                             if len(k_mer_vector) != 64:
#                                 raise Exception('lncRNA 3-mer error')
#                             for number in k_mer_vector:
#                                 node.attributes_vector.append(float(number))
#                             for i in range(49):
#                                 node.attributes_vector.append(0)
#                         if node_type == 'protein':
#                             if len(k_mer_vector) != 49:
#                                 raise Exception('protein 2-mer error')
#                             for i in range(64):
#                                 node.attributes_vector.append(0)
#                             for number in k_mer_vector:
#                                 node.attributes_vector.append(float(number))


def networkx_format_network_generation(interaction_list, negative_interaction_list, lncRNA_list, protein_list):
    edge_list = interaction_list[:]
    edge_list.extend(negative_interaction_list)
    node_list = lncRNA_list[:]
    node_list.extend(protein_list)

    G = nx.Graph()
    for node in node_list:
        G.add_node(node.serial_number)
    for edge in edge_list:
        G.add_edge(edge.lncRNA.serial_number, edge.protein.serial_number)
    print('number of nodes in graph: ', G.number_of_nodes(), 'number of edges in graph: ', G.number_of_edges())
    print(f'number of connected componet : {len(list(nx.connected_components(G)))}')

    del node_list, edge_list
    gc.collect()
    return G


def output_edgelist_file(G, output_path):
    if not osp.exists(output_path):
        os.makedirs(output_path)
        print(f'创建了文件夹：{output_path}')
    output_path += '/bipartite_graph.edgelist'
    if osp.exists(output_path):
        print('edgelist file already exist, rewrite or not? y/n')
        rewrite = input()
        if rewrite == 'n':
            exit()
    nx.write_edgelist(G, path=output_path)


def delete_interaction_from_lncRNA_protein(lncRNA, protein):
    for index in range(len(lncRNA.interaction_list)):
        interaction = lncRNA.interaction_list[index]
        if interaction.protein.serial_number == protein.serial_number:
            del lncRNA.interaction_list[index]
            break
        if index == len(lncRNA.interaction_list)-1:
            raise Exception(f'{lncRNA.name} do not have {lncRNA.name}-{protein.name} we want to delete')
    for index in range(len(protein.interaction_list)):
        interaction = protein.interaction_list[index]
        if interaction.lncRNA.serial_number == lncRNA.serial_number:
            del protein.interaction_list[index]
            break
        if index == len(protein.interaction_list)-1:
            raise Exception(f'{protein.name} do not have {lncRNA.name}-{protein.name} we want to delete')


def reduce_dataset_mentainConnected(G, ratio, list_interaction, list_negativeInteraction, list_lncRNA, list_protein):
    print(f'reduce the dataset to its {ratio}')
    # 确定要删减的数据的数量
    len_list_interaction = len(list_interaction)
    len_list_negativeInteraction = len(list_negativeInteraction)
    num_delete_interaction = int(len(list_interaction) - (len(list_interaction) * ratio))
    num_delete_negativeInteraction = int(len(list_negativeInteraction) - (len(list_negativeInteraction) * ratio))

    # 创建从serial_number到索引的字典，方便找到lncRNA, protein
    dict_serialNumber_listIndex_lncRNA = nodeSerialNumber_listIndex_dict_generation(list_lncRNA)
    dict_serialNumber_listIndex_protein = nodeSerialNumber_listIndex_dict_generation(list_protein)

    # 打乱
    random.shuffle(list_interaction)
    random.shuffle(list_negativeInteraction)

    # 从list_interaction中删除
    num_deleted_positive = 0
    num_deleted_negative = 0
    delete_positive = True
    count = 0
    while num_deleted_positive + num_deleted_negative < num_delete_interaction + num_delete_negativeInteraction:
        # 挑选正例来删除
        delete_positive = not delete_positive
        if delete_positive == True and num_deleted_positive < num_delete_interaction:
            index_interaction = random.sample(range(0, len_list_interaction - num_deleted_positive), 1)[0]
            interaction = list_interaction[index_interaction]
            # 定位到interaction对应的lncRNA和protein
            index_lncRNA = dict_serialNumber_listIndex_lncRNA[interaction.lncRNA.serial_number]
            index_protein = dict_serialNumber_listIndex_protein[interaction.protein.serial_number]
            lncRNA = list_lncRNA[index_lncRNA]
            protein = list_protein[index_protein]
            # 判断删除这个相互作用，会不会使数据集的二部图成为非连通图
            G_temp = G.copy()
            e = (lncRNA.serial_number, protein.serial_number)
            G_temp.remove_edge(*e)
            num_connectedComponent_G_temp = len(list(nx.connected_components(G_temp)))
            if num_connectedComponent_G_temp == 1:
                # 删除这个条边，二部图还是一个连通分量
                delete_interaction_from_lncRNA_protein(lncRNA, protein)
                G = G_temp
                del list_interaction[index_interaction]
                num_deleted_positive += 1
            elif num_connectedComponent_G_temp == 2 and len(lncRNA.interaction_list) == 1:
                # 删除这条边，二部图变成两个连通分量，其中小的那个只包含一个孤立的lncRNA
                delete_interaction_from_lncRNA_protein(lncRNA, protein)
                G_temp.remove_node(lncRNA.serial_number)
                G = G_temp
                del list_interaction[index_interaction]
                num_deleted_positive += 1
            elif num_connectedComponent_G_temp == 2 and len(protein.interaction_list) == 1:
                # 删除这条边，二部图变成两个连通分量，其中小的那个只包含一个孤立的protein
                delete_interaction_from_lncRNA_protein(lncRNA, protein)
                G_temp.remove_node(protein.serial_number)
                G = G_temp
                del list_interaction[index_interaction]
                num_deleted_positive += 1
            # 回收垃圾
            gc.collect()
        elif delete_positive == False and num_deleted_negative < num_delete_negativeInteraction:
            # 挑选负例来删除
            index_interaction = random.sample(range(0, len_list_negativeInteraction - num_deleted_negative), 1)[0]
            interaction = list_negativeInteraction[index_interaction]
            # 定位到interaction对应的lncRNA和protein
            index_lncRNA = dict_serialNumber_listIndex_lncRNA[interaction.lncRNA.serial_number]
            index_protein = dict_serialNumber_listIndex_protein[interaction.protein.serial_number]
            lncRNA = list_lncRNA[index_lncRNA]
            protein = list_protein[index_protein]
            # 判断删除这个相互作用，会不会使数据集的二部图成为非连通图
            G_temp = G.copy()
            e = (lncRNA.serial_number, protein.serial_number)
            G_temp.remove_edge(*e)
            num_connectedComponent_G_temp = len(list(nx.connected_components(G_temp)))
            if num_connectedComponent_G_temp == 1:
                # 删除这个条边，二部图还是一个连通分量
                delete_interaction_from_lncRNA_protein(lncRNA, protein)
                G = G_temp
                del list_negativeInteraction[index_interaction]
                num_deleted_negative += 1
            elif num_connectedComponent_G_temp == 2 and len(lncRNA.interaction_list) == 1:
                # 删除这条边，二部图变成两个连通分量，其中小的那个只包含一个孤立的lncRNA
                # print('删除这条边，二部图变成两个连通分量，其中小的那个只包含一个孤立的lncRNA')
                delete_interaction_from_lncRNA_protein(lncRNA, protein)
                G_temp.remove_node(lncRNA.serial_number)
                G = G_temp
                del list_negativeInteraction[index_interaction]
                num_deleted_negative += 1
            elif num_connectedComponent_G_temp == 2 and len(protein.interaction_list) == 1:
                # 删除这条边，二部图变成两个连通分量，其中小的那个只包含一个孤立的protein
                # print('删除这条边，二部图变成两个连通分量，其中小的那个只包含一个孤立的protein')
                delete_interaction_from_lncRNA_protein(lncRNA, protein)
                G_temp.remove_node(protein.serial_number)
                G = G_temp
                del list_negativeInteraction[index_interaction]
                num_deleted_negative += 1
            # 回收垃圾
            gc.collect()

        if count %100 == 0:
            print(f'number of positive samples need to be deleted: {num_delete_interaction-num_deleted_positive}，number of negative samples need to be deleted: {num_delete_negativeInteraction-num_deleted_negative}')
        count += 1
    print('reduce process done')

    # # 从list_lncRNA中删除掉，已经在二部图中被删除的lncRNA
    # num_deleted_lncRNA = 0
    # num_total_lncRNA = len(list_lncRNA)
    # index_list_lncRNA = 0
    # while index_list_lncRNA < num_total_lncRNA - num_deleted_lncRNA:
    #     if len(list_lncRNA[index_list_lncRNA].interaction_list) == 0:
    #         # 这个lncRNA已经不在二部图里，删除掉
    #         del list_lncRNA[index_list_lncRNA]
    #         num_deleted_lncRNA += 1
    #     else:
    #         index_list_lncRNA += 1
    # # 从list_lncRNA中删除掉，已经在二部图中被删除的protein
    # num_deleted_protein = 0
    # num_total_protein = len(list_protein)
    # index_list_protein = 0
    # while index_list_protein < num_total_protein - num_deleted_protein:
    #     if len(list_protein[index_list_protein].interaction_list) == 0:
    #         # 这个protein已经不在二部图里，删除掉
    #         del list_protein[index_list_protein]
    #         num_deleted_protein += 1
    #     else:
    #         index_list_protein += 1

    # list_interaction_reduced, list_negativeInteraction_reduced, lncRNA_list, protein_list = list_interaction, list_negativeInteraction, list_lncRNA, list_protein

    # # 根据删除过后的list_interaction_reduced, list_negativeInteraction_reduced重新筛选lncRNA_list, protein_list
    # list_serialNumber_lncRNA = []
    # list_serialNumber_protein = []

    # for interaction in list_interaction_reduced:
    #     list_serialNumber_lncRNA.append(interaction.lncRNA.serial_number)
    #     list_serialNumber_protein.append(interaction.protein.serial_number)
    # for interaction in list_negativeInteraction_reduced:
    #     list_serialNumber_lncRNA.append(interaction.lncRNA.serial_number)
    #     list_serialNumber_protein.append(interaction.protein.serial_number)

    # list_serialNumber_lncRNA_reduced = list(set(list_serialNumber_lncRNA))
    # list_serialNumber_protein_reduced = list(set(list_serialNumber_protein))

    # dict_serialNumber_listIndex_lncRNA = nodeSerialNumber_listIndex_dict_generation(lncRNA_list)
    # dict_serialNumber_listIndex_protein = nodeSerialNumber_listIndex_dict_generation(protein_list)

    # list_lncRNA_reduced = []
    # for serial_number in list_serialNumber_lncRNA_reduced:
    #     index = dict_serialNumber_listIndex_lncRNA[serial_number]
    #     list_lncRNA_reduced.append(lncRNA_list[index])
    # list_protein_reduced = []
    # for serial_number in list_serialNumber_protein_reduced:
    #     index = dict_serialNumber_listIndex_protein[serial_number]
    #     list_protein_reduced.append(protein_list[index])

    # # 重建列表的依赖关系
    # list_interaction_reduced, list_negativeInteraction_reduced, list_lncRNA_reduced, list_protein_reduced = reset_basic_data(list_interaction_reduced, list_negativeInteraction_reduced, list_lncRNA_reduced, list_protein_reduced)

    # return list_interaction_reduced, list_negativeInteraction_reduced, list_lncRNA_reduced, list_protein_reduced


# def output_intermediate_products(project_name, interaction_list, negative_interaction_list, lncRNA_list, protein_list):
#     # 消除相互包含关系，防止递归
#     for lncRNA in lncRNA_list:
#         lncRNA.interaction_list = []
#     for protein in protein_list:
#         protein.interaction_list = []

#     output_path = f'data/intermediate_products/{project_name}'
#     if not osp.exists(path=output_path):
#         os.makedirs(output_path)
#         print(f'创建了文件夹：{output_path}')
#     else:
#         raise Exception("corresponding intermediate products exists")

#     print(f'向{output_path}中输出了中间产物')
#     interaction_list_path = f'data/intermediate_products/{project_name}/interaction_list.txt'
#     negative_interaction_list_path = f'data/intermediate_products/{project_name}/negative_interaction_list.txt'
#     lncRNA_list_path = f'data/intermediate_products/{project_name}/lncRNA_list.txt'
#     protein_list_path = f'data/intermediate_products/{project_name}/protein_list.txt'

#     if not osp.exists(path=interaction_list_path):
#         with open(file=interaction_list_path, mode='wb') as f:
#             pickle.dump(interaction_list, f)

#     if not osp.exists(path=negative_interaction_list_path):
#         with open(file=negative_interaction_list_path, mode='wb') as f:
#             pickle.dump(negative_interaction_list, f)

#     if not osp.exists(path=lncRNA_list_path):
#         with open(file=lncRNA_list_path, mode='wb') as f:
#             pickle.dump(lncRNA_list, f)

#     if not osp.exists(path=protein_list_path):
#         with open(file=protein_list_path, mode='wb') as f:
#             pickle.dump(protein_list, f)


def output_set_interactionKey(path:str, set_interactionKey:set):
    print(f'输出了：{path}')
    with open(path, mode='w') as f:
        for interactionKey in set_interactionKey:
            f.write(f'{interactionKey[0]},{interactionKey[1]}\n')


def generate_training_and_testing():
    global set_interactionKey, set_negativeInteractionKey
    
    # 把set_interactionKey和set_negativeInteractionKey分5份
    list_set_interactionKey = [set(), set(), set(), set(), set()]
    list_set_negativeInteractionKey = [set(), set(), set(), set(), set()]
    count = 0
    while len(set_interactionKey) > 0:
        list_set_interactionKey[count % 5].add(set_interactionKey.pop())
        count += 1
    count = 0
    while len(set_negativeInteractionKey) > 0:
        list_set_negativeInteractionKey[count % 5].add(set_negativeInteractionKey.pop())
        count += 1
    
    # 每次那四份组成训练集，另一份是测试集
    for i in range(5):
        set_interactionKey_train = set()
        set_negativeInteractionKey_train = set()
        set_interactionKey_test = set()
        set_negativeInteractionKey_test = set()
        for j in range(5):
            if i == j:
                set_interactionKey_test.update(list_set_interactionKey[j])
                set_negativeInteractionKey_test.update(list_set_negativeInteractionKey[j])
            else:
                set_interactionKey_train.update(list_set_interactionKey[j])
                set_negativeInteractionKey_train.update(list_set_negativeInteractionKey[j])
        if args.output == 1:
            output_set_interactionKey(path_set_allInteractionKey+f'/set_interactionKey_test_{i}', set_interactionKey_test)
            output_set_interactionKey(path_set_allInteractionKey+f'/set_negativeInteractionKey_test_{i}', set_negativeInteractionKey_test)
            output_set_interactionKey(path_set_allInteractionKey+f'/set_interactionKey_train_{i}', set_interactionKey_train)
            output_set_interactionKey(path_set_allInteractionKey+f'/set_negativeInteractionKey_train_{i}', set_negativeInteractionKey_train)
        # 为训练集生成对应的图
        generate_G_training(set_interactionKey_test, set_negativeInteractionKey_test, i)


def generate_G_training(set_interactionKey_test, set_negativeInteraction_test, fold_number):
    global G
    G_whole_temp = copy.deepcopy(G)
    for interactionKey in set_interactionKey_test:
        G_whole_temp.remove_edge(*interactionKey)
    for interactionKey in set_negativeInteraction_test:
        G_whole_temp.remove_edge(*interactionKey)
    G_training = G_whole_temp
    print(f'{fold_number} fold training dataset graph : number of nodes = {G_training.number_of_nodes()}, number of edges = {G_training.number_of_edges()}')
    print(f'number of connected components = {len(list(nx.connected_components(G_training)))}')
    if args.output == 1:
        output_edgelist_file(G_training, f'data/graph/{args.projectName}/training_{fold_number}')


def update_lncRNA_list_and_protein_list(list_interaction:list, list_negativeInteraction:list):
    set_lncRNA = set()
    set_protein = set()
    for interaction in list_interaction:
        set_lncRNA.add(interaction.lncRNA)
        set_protein.add(interaction.protein)
    for interaction in list_negativeInteraction:
        set_lncRNA.add(interaction.lncRNA)
        set_protein.add(interaction.protein)
    print('更新后')
    print(f'number of lncRNAs = {len(set_lncRNA)}, number of proteins = {len(set_protein)}')
    return list(set_lncRNA), list(set_protein)


def return_set_interactionKey(list_interaction):
    set_interactionKey = set()
    for interaction in list_interaction:
        set_interactionKey.add(interaction.key)
    return set_interactionKey


def write_interaction_database_reduced(path:str, list_interaction:list, list_negativeInteraction:list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['RNA names', 'Protein names', 'Labels'])
    for interaction in list_interaction:
        ws.append([interaction.lncRNA.name, interaction.protein.name, 1])
    for negativeInteraction in list_negativeInteraction:
        ws.append([negativeInteraction.lncRNA.name, negativeInteraction.protein.name, 0])
    print(f'输出了: {path}')
    wb.save(path)
    exit(0)


def copy_kmer_data():
    if args.reduce == 1:
        if args.output == 1:
            # 复制一份lncRNA 3mer数据
            path_lncRNA_3_mer = f'data/lncRNA_3_mer/{args.interactionDatasetName}_{args.reduceRatio}'
            if not osp.exists(path_lncRNA_3_mer):
                os.makedirs(path_lncRNA_3_mer)
                print(f'创建了文件夹：{path_lncRNA_3_mer}')
            os.system(f'copy data/lncRNA_3_mer/{args.interactionDatasetName}/lncRNA_3_mer.txt {path_lncRNA_3_mer}/lncRNA_3_mer.txt')
            # 复制一份protein 2mer数据
            path_protein_2_mer = f'data/protein_2_mer/{args.interactionDatasetName}_{args.reduceRatio}'
            if not osp.exists(path_protein_2_mer):
                os.makedirs(path_protein_2_mer)
                print(f'创建了文件夹：{path_protein_2_mer}')
            os.system(f'copy data/protein_2_mer/{args.interactionDatasetName}/protein_2_mer.txt {path_protein_2_mer}/protein_2_mer.txt')
    else:
        raise Exception("this only be called when you reduce dataset")


if __name__ == '__main__':
    print('\n' + 'start' + '\n')

    #参数
    args = parse_args()

    # 正负样本读入或生成
    interaction_dataset_path = 'data/source_database_data/'+ args.interactionDatasetName + '.xlsx'
    interaction_list, negative_interaction_list,lncRNA_list, protein_list, lncRNA_name_index_dict, protein_name_index_dict, set_interactionKey, \
        set_negativeInteractionKey = read_interaction_dataset(dataset_path=interaction_dataset_path, dataset_name=args.interactionDatasetName)
    
    if args.createBalanceDataset == 1:
        negative_interaction_generation() # 生成负样本

    print(f'number of lncRNA: {len(lncRNA_list)}, number of protein: {len(protein_list)}, number of node: {len(lncRNA_list) + len(protein_list)}')
    print(f'number of positive samples: {len(interaction_list)}, number of negative samples: {len(negative_interaction_list)}, number of edges: {len(interaction_list) + len(negative_interaction_list)}')

    # 缩小数据集
    if (args.interactionDatasetName == 'NPInter2' or args.interactionDatasetName == 'NPInter2_0.25') and args.reduce == 1 :
        # 缩小数据集，并且node2vec结果已经准备好
        # 先生成缩小前的edgelist格式的网络
        G = networkx_format_network_generation(interaction_list, negative_interaction_list, lncRNA_list, protein_list)
        # 根据比例缩小数据集
        reduce_dataset_mentainConnected(G, float(args.reduceRatio), interaction_list, negative_interaction_list, lncRNA_list, protein_list)
        lncRNA_list, protein_list = update_lncRNA_list_and_protein_list(interaction_list, negative_interaction_list)
        set_interactionKey = return_set_interactionKey(interaction_list)
        set_negativeInteractionKey = return_set_interactionKey(negative_interaction_list)
        # 输出缩小后的网络
        G = networkx_format_network_generation(interaction_list, negative_interaction_list, lncRNA_list, protein_list)
        graph_output_path = f'data/graph/{args.projectName}'
        if(args.output == 1):
            # 把削减后的相互作用数据集存起来
            path_interaction_database_reduced = f'data/source_database_data/{args.interactionDatasetName}_{args.reduceRatio}.xlsx'
            write_interaction_database_reduced(path_interaction_database_reduced, interaction_list, negative_interaction_list)
            # 把k-mer的数据也复制
            # copy_kmer_data()
            output_edgelist_file(G, graph_output_path)
            # 把生成的负样本存起来
            path_set_allInteractionKey = f'data/set_allInteractionKey/{args.projectName}'
            if not osp.exists(path_set_allInteractionKey):
                print(f'创建了文件夹：{path_set_allInteractionKey}')
                os.makedirs(path_set_allInteractionKey)
            output_set_interactionKey(path_set_allInteractionKey+'/set_negativeInteractionKey_all', set_negativeInteractionKey)
        # 生成训练集-测试集
        generate_training_and_testing()
            # 输出中间产物
        #     output_intermediate_products(args.projectName, args.node2vecWindowSize, interaction_list, negative_interaction_list, lncRNA_list, protein_list)
    elif args.interactionDatasetName != 'NPInter2' and args.reduce == 1 :
        raise Exception("temporary do not support reduce interaction dataset other than NPInter2")
    else:
        # 不缩小数据集
        # 生成edgelist格式的网络并保存
        G = networkx_format_network_generation(interaction_list, negative_interaction_list, lncRNA_list, protein_list)
        graph_output_path = f'data/graph/{args.projectName}'
        if(args.output == 1):
            output_edgelist_file(G, graph_output_path)
            # 输出中间产物
            # output_intermediate_products(args.projectName, interaction_list, negative_interaction_list, lncRNA_list, protein_list)
            path_set_allInteractionKey = f'data/set_allInteractionKey/{args.projectName}'
            if not osp.exists(path_set_allInteractionKey):
                print(f'创建了文件夹：{path_set_allInteractionKey}')
                os.makedirs(path_set_allInteractionKey)
            output_set_interactionKey(path_set_allInteractionKey+'/set_negativeInteractionKey_all', set_negativeInteractionKey)
        # 生成训练集-测试集
        generate_training_and_testing()

    # 创建保存node2vec结果的文件夹
    path_node2vec_result = f'data\\node2vec_result\\{args.projectName}'
    if not osp.exists(path_node2vec_result) and args.output == 1:
        os.makedirs(path_node2vec_result)
        print(f'创建了文件夹: {path_node2vec_result}')
        for i in range(5):
            path_node2vec_per_fold = path_node2vec_result + f'/training_{i}'
            os.makedirs(path_node2vec_per_fold)
            print(f'创建了文件夹: {path_node2vec_per_fold}')
    
    print('\n' + 'exit' + '\n')